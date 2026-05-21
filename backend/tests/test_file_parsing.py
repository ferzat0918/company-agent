import base64
import io
import zipfile
import pytest
from langchain_core.messages import HumanMessage
from src.chat_models import (
    extract_pdf_text_from_base64,
    extract_docx_text_from_bytes,
    extract_xlsx_text_from_bytes,
    extract_text_from_bytes,
    preprocess_messages,
)


def test_extract_text_from_bytes():
    # UTF-8
    assert extract_text_from_bytes(b"hello world") == "hello world"
    # GBK fallback
    assert extract_text_from_bytes("你好 gbk".encode("gbk")) == "你好 gbk"


def test_extract_docx_text_from_bytes():
    # Create an in-memory docx zip
    docx_io = io.BytesIO()
    with zipfile.ZipFile(docx_io, "w") as docx:
        # docx needs word/document.xml with paragraph elements
        xml_content = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
            <w:body>
                <w:p>
                    <w:r>
                        <w:t>Hello Word Document Paragraph 1</w:t>
                    </w:r>
                </w:p>
                <w:p>
                    <w:r>
                        <w:t>Second paragraph text</w:t>
                    </w:r>
                </w:p>
            </w:body>
        </w:document>
        """
        docx.writestr("word/document.xml", xml_content)
    
    extracted = extract_docx_text_from_bytes(docx_io.getvalue())
    assert "Hello Word Document Paragraph 1" in extracted
    assert "Second paragraph text" in extracted


def test_extract_xlsx_text_from_bytes():
    # Use openpyxl to write a mock xlsx
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Name"
        ws["B1"] = "Score"
        ws["A2"] = "Alice"
        ws["B2"] = 95
        
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        
        extracted = extract_xlsx_text_from_bytes(xlsx_io.getvalue())
        assert "[工作表: TestSheet]" in extracted
        assert "Name | Score" in extracted
        assert "Alice | 95" in extracted
    except ImportError:
        # If openpyxl is not installed in the testing environment, it should fail with import error string
        assert "未安装 openpyxl" in extract_xlsx_text_from_bytes(b"")


def test_preprocess_messages():
    # Test message text document parsing
    raw_text = "line 1\nline 2"
    raw_base64 = base64.b64encode(raw_text.encode("utf-8")).decode("utf-8")
    
    msg = HumanMessage(content=[
        {"type": "text", "text": "Please analyze this file:"},
        {
            "type": "file",
            "mimeType": "text/plain",
            "data": raw_base64,
            "metadata": {"filename": "info.txt"}
        },
        {
            "type": "image",
            "mimeType": "image/png",
            "data": "dummy-base64",
            "metadata": {"name": "chart.png"}
        }
    ])
    
    preprocessed = preprocess_messages([msg])
    assert len(preprocessed) == 1
    new_msg = preprocessed[0]
    assert isinstance(new_msg, HumanMessage)
    
    # Check that text block is kept
    assert new_msg.content[0]["text"] == "Please analyze this file:"
    
    # Check that plain text file was parsed
    assert "[文本文件附件: info.txt 的内容开始]" in new_msg.content[1]["text"]
    assert "line 1\nline 2" in new_msg.content[1]["text"]
    
    # Check that image block was converted to text block indicating visual unsupported
    assert "[图片附件: chart.png" in new_msg.content[2]["text"]
