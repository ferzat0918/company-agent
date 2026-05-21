"""ChatDeepSeek with reasoning_content preservation and document/multimodal preprocessing.

DeepSeek's thinking mode requires that when the model performs tool calls,
the reasoning_content from that assistant message must be passed back verbatim
in all subsequent requests. LangChain's message serialization
(_convert_message_to_dict) strips this field from AIMessage.additional_kwargs,
causing a 400 error. This subclass restores it.

Additionally, this class preprocesses message inputs, transparently converting
multimodal blocks (images, PDFs, Word documents, Excel sheets, and text files)
into plain text context before calling the text-only DeepSeek API.
"""
import base64
import io
from collections.abc import Sequence
from typing import Any

from langchain_core.language_models import LanguageModelInput
from langchain_core.messages import AIMessage, HumanMessage
from langchain_deepseek import ChatDeepSeek


def extract_pdf_text_from_base64(base64_str: str) -> str:
    """Extract text from base64-encoded PDF bytes."""
    try:
        import pypdf
        
        pdf_bytes = base64.b64decode(base64_str)
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        
        text = text.strip()
        if not text:
            return "(该PDF文件没有提取出任何可读文本，可能是扫描件或纯图片PDF)"
        return text
    except ImportError:
        return "(解析失败: 后端未安装 pypdf 依赖库，请联系系统管理员安装 pypdf)"
    except Exception as e:
        return f"(解析PDF时出错: {str(e)})"


def extract_docx_text_from_bytes(docx_bytes: bytes) -> str:
    """Extract text from docx bytes using standard libraries (no external dependencies)."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        
        with zipfile.ZipFile(io.BytesIO(docx_bytes)) as docx:
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            paragraphs = []
            for el in root.iter():
                if el.tag.endswith('}p'):
                    texts = []
                    for text_el in el.findall('.//w:t', ns):
                        if text_el.text:
                            texts.append(text_el.text)
                    if texts:
                        paragraphs.append("".join(texts))
            
            text = "\n".join(paragraphs).strip()
            if not text:
                return "(该Word文件内无文本内容)"
            return text
    except Exception as e:
        return f"(解析Word时出错: {str(e)})"


def extract_xlsx_text_from_bytes(xlsx_bytes: bytes) -> str:
    """Extract text from xlsx bytes using openpyxl, formatting worksheets as tabular text."""
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        sheets_text = []
        for name in wb.sheetnames:
            sheet = wb[name]
            sheet_lines = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    sheet_lines.append(row_str)
            if sheet_lines:
                sheets_text.append(f"[工作表: {name}]\n" + "\n".join(sheet_lines))
        
        text = "\n\n".join(sheets_text).strip()
        if not text:
            return "(该Excel文件内无数据内容)"
        return text
    except ImportError:
        return "(解析失败: 后端未安装 openpyxl 依赖库，请联系系统管理员安装 openpyxl)"
    except Exception as e:
        return f"(解析Excel时出错: {str(e)})"


def extract_text_from_bytes(text_bytes: bytes) -> str:
    """Extract string from raw text bytes with encoding detection fallbacks."""
    try:
        return text_bytes.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return text_bytes.decode("gbk", errors="ignore")
        except Exception as e:
            return f"(解析文本文件失败: {str(e)})"


def _process_content_list(content_list: list) -> list:
    """Helper to process file/image blocks inside a list content and convert to text."""
    new_content = []
    for block in content_list:
        if isinstance(block, dict):
            block_type = block.get("type")
            if block_type == "file":
                mime_type = block.get("mimeType")
                filename = (
                    block.get("metadata", {}).get("filename")
                    or block.get("metadata", {}).get("name")
                    or "document"
                )
                data_base64 = block.get("data")
                
                if data_base64:
                    try:
                        file_bytes = base64.b64decode(data_base64)
                    except Exception as e:
                        new_content.append({
                            "type": "text",
                            "text": f"\n[文件附件解码失败: {filename} ({str(e)})]\n"
                        })
                        continue

                    # Route parsing based on MIME-type or file extension
                    is_pdf = mime_type == "application/pdf" or filename.lower().endswith(".pdf")
                    is_docx = (
                        mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        or filename.lower().endswith(".docx")
                    )
                    is_xlsx = (
                        mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        or filename.lower().endswith(".xlsx")
                    )
                    is_text = (
                        mime_type in ["text/plain", "text/markdown", "text/csv", "application/json"]
                        or filename.lower().endswith((".txt", ".md", ".csv", ".json", ".xml", ".ini", ".log"))
                    )

                    if is_pdf:
                        extracted_text = extract_pdf_text_from_base64(data_base64)
                        new_content.append({
                            "type": "text",
                            "text": f"\n[PDF文件附件: {filename} 的文本内容开始]\n{extracted_text}\n[PDF文件附件: {filename} 的文本内容结束]\n"
                        })
                    elif is_docx:
                        extracted_text = extract_docx_text_from_bytes(file_bytes)
                        new_content.append({
                            "type": "text",
                            "text": f"\n[Word文件附件: {filename} 的文本内容开始]\n{extracted_text}\n[Word文件附件: {filename} 的文本内容结束]\n"
                        })
                    elif is_xlsx:
                        extracted_text = extract_xlsx_text_from_bytes(file_bytes)
                        new_content.append({
                            "type": "text",
                            "text": f"\n[Excel表格附件: {filename} 的数据内容开始]\n{extracted_text}\n[Excel表格附件: {filename} 的数据内容结束]\n"
                        })
                    elif is_text:
                        extracted_text = extract_text_from_bytes(file_bytes)
                        new_content.append({
                            "type": "text",
                            "text": f"\n[文本文件附件: {filename} 的内容开始]\n{extracted_text}\n[文本文件附件: {filename} 的内容结束]\n"
                        })
                    else:
                        new_content.append({
                            "type": "text",
                            "text": f"\n[文件附件: {filename} (类型: {mime_type}, 暂不支持自动提取内容)]\n"
                        })
                else:
                    new_content.append({
                        "type": "text",
                        "text": f"\n[空文件附件: {filename}]\n"
                    })
            elif block_type == "image":
                filename = (
                    block.get("metadata", {}).get("name")
                    or block.get("metadata", {}).get("filename")
                    or "image.png"
                )
                new_content.append({
                    "type": "text",
                    "text": f"\n[图片附件: {filename} (当前使用的是纯文本模型 DeepSeek，无法直接视觉解析此图片，仅读取了文件名)]\n"
                })
            else:
                new_content.append(block)
        else:
            new_content.append(block)
    return new_content


def preprocess_messages(input_: LanguageModelInput) -> LanguageModelInput:
    """Preprocess LangChain message inputs.
    
    Converts unsupported multimodal content blocks (like 'file' representing PDFs,
    Word, Excel, or Text documents, and text-only unsupported 'image' blocks)
    into standard text blocks readable by DeepSeek.
    """
    try:
        with open("/tmp/preprocess_debug.txt", "a", encoding="utf-8") as f:
            f.write(f"\n--- preprocess_messages called ---\n")
            f.write(f"type(input_): {type(input_)}\n")
            f.write(f"input_ representation: {repr(input_)[:2000]}\n")
    except Exception as e:
        pass

    if not isinstance(input_, Sequence) or isinstance(input_, str):
        return input_
        
    new_messages = []
    for msg in input_:
        if isinstance(msg, HumanMessage) and isinstance(msg.content, list):
            new_content = _process_content_list(msg.content)
            msg = HumanMessage(
                content=new_content,
                additional_kwargs=msg.additional_kwargs,
                response_metadata=msg.response_metadata,
                id=msg.id,
                name=msg.name
            )
        elif isinstance(msg, dict) and (msg.get("type") == "human" or msg.get("role") in ("user", "human")) and isinstance(msg.get("content"), list):
            new_content = _process_content_list(msg["content"])
            msg = dict(msg)
            msg["content"] = new_content
        elif hasattr(msg, "type") and getattr(msg, "type") == "human" and isinstance(getattr(msg, "content", None), list):
            new_content = _process_content_list(msg.content)
            try:
                msg = msg.__class__(
                    content=new_content,
                    additional_kwargs=msg.additional_kwargs,
                    response_metadata=msg.response_metadata,
                    id=msg.id,
                    name=msg.name
                )
            except Exception:
                msg = HumanMessage(
                    content=new_content,
                    additional_kwargs=getattr(msg, "additional_kwargs", {}),
                    response_metadata=getattr(msg, "response_metadata", {}),
                    id=getattr(msg, "id", None),
                    name=getattr(msg, "name", None)
                )
        new_messages.append(msg)
    return new_messages


class ChatDeepSeekThinking(ChatDeepSeek):
    """ChatDeepSeek that preserves reasoning_content for multi-turn conversations and processes file/image uploads."""

    def _get_request_payload(
        self,
        input_: LanguageModelInput,
        *,
        stop: list[str] | None = None,
        **kwargs: Any,
    ) -> dict:
        # Preprocess multimodal/file inputs into text before serialization
        preprocessed_input = preprocess_messages(input_)
        payload = super()._get_request_payload(preprocessed_input, stop=stop, **kwargs)

        if isinstance(preprocessed_input, Sequence) and not isinstance(preprocessed_input, str):
            for i, msg in enumerate(preprocessed_input):
                if i >= len(payload["messages"]):
                    break
                if isinstance(msg, AIMessage):
                    reasoning = msg.additional_kwargs.get("reasoning_content")
                    if reasoning:
                        payload["messages"][i]["reasoning_content"] = reasoning

        return payload
